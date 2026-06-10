import csv
import re
import unicodedata
from io import BytesIO, StringIO
from pathlib import Path
from typing import Iterable

from .models import VideoRecord


def normalize_columns(row: dict) -> dict:
    normalized = {}
    for key, value in row.items():
        if key is None:
            continue
        clean_key = str(key).strip().lower().replace(" ", "_")
        normalized[clean_key] = _clean_text(value)
    return normalized


def read_legacy_rows(file_bytes: bytes, filename: str) -> list[dict]:
    suffix = Path(filename).suffix.lower()
    if suffix == ".csv":
        return _read_csv(file_bytes)
    if suffix == ".xlsx":
        return _read_xlsx(file_bytes)
    raise ValueError("Only .csv and .xlsx files are supported.")


def rows_to_video_records(rows: Iterable[dict]) -> list[VideoRecord]:
    records = []
    for row in rows:
        item = normalize_columns(row)

        source_url = item.get("source_url", "")
        title = item.get("title", "")
        filename = item.get("filename", "") or item.get("local_filename", "") or item.get("local_path", "")
        schedule_time = item.get("schedule_time", "") or item.get("scheduled_time", "") or item.get("schedule", "")

        if not any([source_url, title, filename, schedule_time]):
            continue

        records.append(
            VideoRecord(
                source_url=source_url,
                source_video_id=_source_video_id(item, source_url),
                source_page_url=item.get("source_page_url", "") or item.get("source_profile_url", ""),
                title_original=title,
                title_rewrite=item.get("title_rewrite", ""),
                created_at=item.get("created_at", ""),
                local_filename=filename,
                download_status=_infer_download_status(item, filename),
                publish_status=_infer_publish_status(item),
                schedule_time=schedule_time,
                fb_post_id=item.get("fb_post_id", "") or item.get("facebook_video_id", ""),
                error_message=item.get("error_message", ""),
            )
        )
    return records


def _source_video_id(item: dict, source_url: str) -> str:
    explicit = item.get("source_video_id", "").strip()
    if explicit:
        return explicit

    # Legacy CSV files use video_id as a row number, not a stable Facebook id.
    # Derive a stable id only from recognizable Facebook URL shapes.
    if not source_url:
        return ""

    patterns = [
        r"/videos/([^/?#]+)/?",
        r"/watch/\?v=([^&#]+)",
        r"/reel/([^/?#]+)/?",
        r"/share/v/([^/?#]+)/?",
    ]
    for pattern in patterns:
        match = re.search(pattern, source_url)
        if match:
            return match.group(1).strip()
    return ""


def _read_csv(file_bytes: bytes) -> list[dict]:
    text = _decode_csv_text(file_bytes)

    return [normalize_columns(row) for row in csv.DictReader(StringIO(text))]


def _decode_csv_text(file_bytes: bytes) -> str:
    encodings = ["utf-8-sig", "utf-8", "cp1258", "cp1252", "latin-1"]
    candidates = []
    for encoding in encodings:
        try:
            text = file_bytes.decode(encoding)
        except UnicodeDecodeError:
            continue
        candidates.append((_decode_score(text), encoding, text))

    if not candidates:
        raise ValueError("Could not decode CSV file.")

    candidates.sort(reverse=True)
    return candidates[0][2]


def _decode_score(text: str) -> int:
    vietnamese_chars = set(
        "ăâđêôơư"
        "áàảãạấầẩẫậắằẳẵặ"
        "éèẻẽẹếềểễệ"
        "íìỉĩị"
        "óòỏõọốồổỗộớờởỡợ"
        "úùủũụứừửữự"
        "ýỳỷỹỵ"
        "ĂÂĐÊÔƠƯ"
        "ÁÀẢÃẠẤẦẨẪẬẮẰẲẴẶ"
        "ÉÈẺẼẸẾỀỂỄỆ"
        "ÍÌỈĨỊ"
        "ÓÒỎÕỌỐỒỔỖỘỚỜỞỠỢ"
        "ÚÙỦŨỤỨỪỬỮỰ"
        "ÝỲỶỸỴ"
    )
    normalized = unicodedata.normalize("NFC", text)
    score = sum(2 for char in normalized if char in vietnamese_chars)
    score -= normalized.count("Ã") * 3
    score -= normalized.count("Â") * 2
    score -= normalized.count("�") * 10
    return score


def _clean_text(value) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    return unicodedata.normalize("NFC", text)


def _read_xlsx(file_bytes: bytes) -> list[dict]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to import .xlsx files.") from exc

    workbook = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = workbook.active
    rows = sheet.iter_rows(values_only=True)
    headers = next(rows, None)
    if not headers:
        return []

    clean_headers = [str(header).strip() if header is not None else "" for header in headers]
    result = []
    for values in rows:
        row = {
            clean_headers[idx]: values[idx] if idx < len(values) else ""
            for idx in range(len(clean_headers))
            if clean_headers[idx]
        }
        result.append(normalize_columns(row))
    return result


def _infer_download_status(item: dict, filename: str) -> str:
    status = item.get("download_status", "")
    if status:
        return status
    legacy_status = item.get("status", "")
    if legacy_status in {"success", "downloaded", "download_failed"}:
        return "downloaded" if legacy_status == "success" else legacy_status
    if filename:
        return "downloaded"
    return "pending"


def _infer_publish_status(item: dict) -> str:
    status = item.get("publish_status", "")
    if status:
        return status
    legacy_status = item.get("status", "")
    if legacy_status in {"published", "scheduled", "failed"}:
        return "post_failed" if legacy_status == "failed" else legacy_status
    return "pending"
